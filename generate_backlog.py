import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define the complete 64 User Stories with Agile Scrum standards
backlog_data = [
    # Epic 1: Auth & Security (Sprint 1)
    {
        "epic": "1. XÁC THỰC & BẢO MẬT (AUTH)", "code": "US-01",
        "story": "Là Nhân viên, tôi muốn đăng nhập hệ thống bằng Email và Mật khẩu để tôi truy cập vào đúng các chức năng nghiệp vụ của mình.",
        "points": 3, "priority": "Cao",
        "ac": "1. Có form kiểm tra định dạng email hợp lệ.\n2. Mật khẩu nhập vào phải được ẩn.\n3. Khóa tài khoản tạm thời trong 15 phút nếu nhập sai quá 5 lần liên tiếp.",
        "status": "Done", "dev": "Kha", "tester": "Thuận", "sprint": 1, "notes": "Đã tích hợp mã hóa BCrypt"
    },
    {
        "epic": "1. XÁC THỰC & BẢO MẬT (AUTH)", "code": "US-02",
        "story": "Là Người dùng, tôi muốn đăng nhập nhanh thông qua liên kết tài khoản Google OAuth2 để rút ngắn thời gian truy cập.",
        "points": 3, "priority": "Cao",
        "ac": "1. Hiển thị nút 'Đăng nhập bằng Google' ở trang chủ.\n2. Tự động ánh xạ thông tin email Google với tài khoản nhân viên trong DB.\n3. Từ chối đăng nhập nếu email Google chưa được khai báo.",
        "status": "Done", "dev": "Kha", "tester": "Thuận", "sprint": 1, "notes": "Sử dụng Spring Security OAuth2"
    },
    {
        "epic": "1. XÁC THỰC & BẢO MẬT (AUTH)", "code": "US-03",
        "story": "Là Người quản lý, tôi muốn kích hoạt xác thực 2 yếu tố (2FA TOTP) bằng Google Authenticator để gia tăng bảo mật dữ liệu nhạy cảm.",
        "points": 5, "priority": "Trung bình",
        "ac": "1. Hiển thị mã QR và khóa bí mật (Secret Key) khi kích hoạt 2FA.\n2. Yêu cầu nhập mã OTP 6 số từ Google Authenticator để xác nhận bật thành công.\n3. Cho phép tắt 2FA sau khi nhập mật khẩu xác nhận.",
        "status": "Done", "dev": "Kha", "tester": "Thuận", "sprint": 1, "notes": "Sử dụng thư viện Google TOTP"
    },
    {
        "epic": "1. XÁC THỰC & BẢO MẬT (AUTH)", "code": "US-04",
        "story": "Là Người dùng, tôi muốn nhận mã xác thực OTP gửi qua Email khi đăng nhập từ thiết bị lạ để ngăn chặn truy cập trái phép.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Tự động quét IP đăng nhập.\n2. Nếu phát hiện IP lạ, yêu cầu nhập mã OTP gửi qua email đăng ký.\n3. Mã OTP qua email có hiệu lực trong vòng 3 phút.",
        "status": "Done", "dev": "Kha", "tester": "Thuận", "sprint": 1, "notes": "Sử dụng JavaMailSender"
    },
    {
        "epic": "1. XÁC THỰC & BẢO MẬT (AUTH)", "code": "US-05",
        "story": "Là Người dùng quên mật khẩu, tôi muốn yêu cầu đặt lại mật khẩu qua email xác nhận để tôi chủ động lấy lại quyền truy cập.",
        "points": 3, "priority": "Cao",
        "ac": "1. Có trang 'Quên mật khẩu' yêu cầu nhập email.\n2. Gửi link reset mật khẩu chứa token bảo mật dùng 1 lần (hạn dùng 15 phút).\n3. Cho phép nhập mật khẩu mới và xác nhận mật khẩu mới.",
        "status": "Done", "dev": "Dũng", "tester": "Lộc", "sprint": 1, "notes": "Token hết hạn sau 15 phút"
    },
    {
        "epic": "1. XÁC THỰC & BẢO MẬT (AUTH)", "code": "US-06",
        "story": "Là Quản trị viên (Admin), tôi muốn thiết lập phân quyền vai trò chi tiết (RBAC) để giới hạn quyền truy cập các endpoint nghiệp vụ.",
        "points": 5, "priority": "Cao",
        "ac": "1. Định nghĩa các role: ADMIN, MANAGER, CASHIER, KITCHEN, HR, WAREHOUSE.\n2. Giới hạn truy cập API bằng Spring Security `@PreAuthorize`.\n3. Trả về mã lỗi 403 Forbidden nếu người dùng không đủ thẩm quyền.",
        "status": "Done", "dev": "Dũng", "tester": "Lộc", "sprint": 1, "notes": "Phân quyền theo Role-based"
    },
    {
        "epic": "1. XÁC THỰC & BẢO MẬT (AUTH)", "code": "US-07",
        "story": "Là Quản trị viên (Admin), tôi muốn hệ thống tự động ghi nhật ký kiểm toán (Audit Logs) để tôi theo dõi hoạt động nhạy cảm.",
        "points": 3, "priority": "Cao",
        "ac": "1. Ghi nhận: ID người dùng, Hành động, Kiểu thực thể tác động, Chi tiết thay đổi, Địa chỉ IP, Thời điểm.\n2. Lưu vết khi thêm/sửa/xóa hóa đơn, điều chỉnh kho, thăng hạng thẻ khách.\n3. Không cho phép bất kỳ ai sửa hoặc xóa bản ghi trong bảng audit_logs.",
        "status": "Done", "dev": "Dũng", "tester": "Lộc", "sprint": 1, "notes": "Lưu trữ bảo mật, chỉ đọc"
    },

    # Epic 2: POS & Tables (Sprint 1)
    {
        "epic": "2. SƠ ĐỒ BÀN & POS (POS)", "code": "US-08",
        "story": "Là Thu ngân, tôi muốn xem sơ đồ khu vực và bàn ăn thời gian thực để tôi nắm rõ trạng thái trống/đầy và phục vụ khách.",
        "points": 5, "priority": "Cao",
        "ac": "1. Hiển thị danh sách các khu vực (Sân trước, Phòng lạnh, Lầu 1...).\n2. Mỗi bàn ăn hiển thị màu sắc theo trạng thái: Xám (Trống), Đỏ (Đang ăn), Vàng (Đặt trước).\n3. Tự động đồng bộ trạng thái bàn khi có thay đổi từ các máy POS khác.",
        "status": "In Progress", "dev": "Thuận", "tester": "Kha", "sprint": 1, "notes": "Đồng bộ WebSocket"
    },
    {
        "epic": "2. SƠ ĐỒ BÀN & POS (POS)", "code": "US-09",
        "story": "Là Thu ngân, tôi muốn mở phiên làm việc của bàn (Table Session) để bắt đầu gọi món phục vụ khách ngồi tại bàn.",
        "points": 3, "priority": "Cao",
        "ac": "1. Cho phép nhấp vào bàn trống để 'Mở bàn'.\n2. Cho phép nhập số lượng khách ngồi thực tế (guest_count).\n3. Trạng thái bàn chuyển từ EMPTY sang OCCUPIED, tạo bản ghi table_sessions với trạng thái ACTIVE.",
        "status": "In Progress", "dev": "Thuận", "tester": "Kha", "sprint": 1, "notes": "Lưu table_sessions"
    },
    {
        "epic": "2. SƠ ĐỒ BÀN & POS (POS)", "code": "US-10",
        "story": "Là Thu ngân, tôi muốn thêm món ăn vào giỏ hàng bàn ăn để ghi nhận yêu cầu gọi món của khách hàng.",
        "points": 3, "priority": "Cao",
        "ac": "1. Hiển thị thực đơn trực quan phân loại theo Danh mục.\n2. Nhấp vào món ăn để thêm vào giỏ hàng bàn hiện tại.\n3. Tự động tính toán tổng tiền tạm tính của giỏ hàng.",
        "status": "In Progress", "dev": "Lộc", "tester": "Dũng", "sprint": 1, "notes": "Tích hợp giỏ hàng POS"
    },
    {
        "epic": "2. SƠ ĐỒ BÀN & POS (POS)", "code": "US-11",
        "story": "Là Thu ngân, tôi muốn chọn biến thể món ăn (Size, Toppings, Đá, Nóng) để ghi nhận chi tiết khẩu vị khách.",
        "points": 3, "priority": "Cao",
        "ac": "1. Khi chọn món, hiển thị cửa sổ popup cấu hình biến thể sản phẩm.\n2. Cho phép chọn 1 size (Standard, Medium, Large) và nhiều topping ăn kèm.\n3. Giá bán của biến thể và topping tự động cộng dồn vào đơn giá món ăn.",
        "status": "In Progress", "dev": "Dũng", "tester": "Lộc", "sprint": 1, "notes": "Lưu product_variants"
    },
    {
        "epic": "2. SƠ ĐỒ BÀN & POS (POS)", "code": "US-12",
        "story": "Là Thu ngân, tôi muốn gửi yêu cầu làm món xuống Bếp để đầu bếp chuẩn bị chế biến đúng theo thứ tự gọi món.",
        "points": 5, "priority": "Cao",
        "ac": "1. Có nút 'Gửi bếp' trên giỏ hàng POS.\n2. Sau khi gửi, khởi tạo bản ghi trong bảng orders (status = PENDING).\n3. Đẩy tin nhắn thông báo WebSocket chứa danh sách món cần làm sang màn hình hiển thị bếp KDS.",
        "status": "In Progress", "dev": "Dũng", "tester": "Lộc", "sprint": 1, "notes": "Đẩy WebSocket sang KDS"
    },
    {
        "epic": "2. SƠ ĐỒ BÀN & POS (POS)", "code": "US-13",
        "story": "Là Thu ngân, tôi muốn tách hóa đơn thanh toán (Split Bill) khi một nhóm khách muốn tự thanh toán riêng phần của họ.",
        "points": 5, "priority": "Trung bình",
        "ac": "1. Cho phép chọn các món ăn cụ thể trong giỏ hàng hiện tại.\n2. Chuyển các món đã chọn sang một hóa đơn mới riêng biệt.\n3. Giữ nguyên các món còn lại ở hóa đơn gốc tại bàn ăn.",
        "status": "In Progress", "dev": "Kha", "tester": "Thuận", "sprint": 1, "notes": "Tách hóa đơn POS"
    },
    {
        "epic": "2. SƠ ĐỒ BÀN & POS (POS)", "code": "US-14",
        "story": "Là Thu ngân, tôi muốn ghép hóa đơn nhiều bàn (Merge Bill) khi các bàn đi chung nhóm muốn gộp tiền trả một thể.",
        "points": 5, "priority": "Trung bình",
        "ac": "1. Cho phép chọn bàn nguồn và bàn đích cần gộp.\n2. Chuyển toàn bộ món từ bàn nguồn sang bàn đích.\n3. Tự động hủy phiên ăn ở bàn nguồn, chuyển bàn nguồn về trạng thái trống (EMPTY).",
        "status": "In Progress", "dev": "Lộc", "tester": "Dũng", "sprint": 1, "notes": "Gộp hóa đơn POS"
    },
    {
        "epic": "2. SƠ ĐỒ BÀN & POS (POS)", "code": "US-15",
        "story": "Là Thu ngân, tôi muốn xuất mã QR VNPay để khách hàng quét thanh toán hóa đơn nhanh chóng và hạn chế sai số tiền mặt.",
        "points": 5, "priority": "Cao",
        "ac": "1. Khi chọn thanh toán VNPay, gọi API VNPay Sandbox sinh chuỗi URL QR Code thanh toán.\n2. Hiển thị mã QR lên màn hình POS.\n3. Lắng nghe phản hồi giao dịch qua IPN endpoint để tự động xác nhận hóa đơn đã thanh toán.",
        "status": "In Progress", "dev": "Kha", "tester": "Thuận", "sprint": 1, "notes": "Tích hợp SDK VNPay Sandbox"
    },

    # Epic 3: KDS Display (Sprint 2)
    {
        "epic": "3. HỆ THỐNG HIỂN THỊ BẾP (KDS)", "code": "US-16",
        "story": "Là Đầu bếp, tôi muốn xem danh sách các món ăn cần chế biến hiển thị thời gian thực theo thứ tự thời gian đặt để chuẩn bị kịp thời.",
        "points": 5, "priority": "Cao",
        "ac": "1. Màn hình KDS tự động cập nhật món ăn mới ngay khi thu ngân nhấn 'Gửi bếp' mà không cần tải lại trang.\n2. Hiển thị chi tiết: Tên món, Biến thể, Số lượng, Ghi chú chế biến, Thời gian chờ (đếm ngược).\n3. Sắp xếp món theo thứ tự thời gian gọi món từ cũ đến mới.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 2, "notes": "Kết nối WebSocket KDS"
    },
    {
        "epic": "3. HỆ THỐNG HIỂN THỊ BẾP (KDS)", "code": "US-17",
        "story": "Là Đầu bếp, tôi muốn cập nhật trạng thái món ăn sang 'Đang làm' (COOKING) để thu ngân biết món ăn đang được chế biến.",
        "points": 2, "priority": "Cao",
        "ac": "1. Nhấp vào thẻ món ăn trên KDS để đổi trạng thái sang COOKING.\n2. Đồng bộ trạng thái về màn hình POS của thu ngân thời gian thực.\n3. Cập nhật trạng thái trong bảng `order_details.status = 'COOKING'`.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 2, "notes": "Cập nhật order_details"
    },
    {
        "epic": "3. HỆ THỐNG HIỂN THỊ BẾP (KDS)", "code": "US-18",
        "story": "Là Đầu bếp, tôi muốn cập nhật trạng thái món ăn sang 'Đã xong' (READY) để phục vụ mang món ăn lên cho khách hàng.",
        "points": 2, "priority": "Cao",
        "ac": "1. Nhấp vào nút 'Hoàn thành' trên thẻ món ở KDS.\n2. Đổi trạng thái sang READY, tự động đẩy thẻ món ăn vào danh sách chờ bưng.\n3. Bắn WebSocket thông báo pop-up lên màn hình POS báo phục vụ mang món ra bàn ăn.",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 2, "notes": "Bắn WebSocket READY"
    },
    {
        "epic": "3. HỆ THỐNG HIỂN THỊ BẾP (KDS)", "code": "US-19",
        "story": "Là Người quản lý, tôi muốn hệ thống hiển thị cảnh báo đỏ khi một món ăn bị chế biến trễ quá 15 phút để tôi kịp thời điều phối bếp.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Hệ thống đếm thời gian từ lúc order ở trạng thái PENDING.\n2. Nếu sau 15 phút chưa đổi sang READY, đổi màu thẻ món ăn trên KDS thành màu đỏ nhấp nháy.\n3. Gửi cảnh báo hiệu suất lên dashboard quản trị.",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 2, "notes": "Hệ thống đếm thời gian thực"
    },
    {
        "epic": "3. HỆ THỐNG HIỂN THỊ BẾP (KDS)", "code": "US-20",
        "story": "Là Thu ngân, tôi muốn đánh dấu món ăn đã được phục vụ (SERVED) cho khách tại bàn để tôi chốt số lượng món thực tế đã dùng.",
        "points": 2, "priority": "Cao",
        "ac": "1. Trên giỏ hàng POS hiển thị danh sách món READY.\n2. Cho phép nhấp chọn 'Đã phục vụ' (SERVED) khi phục vụ bưng món xong.\n3. Cập nhật `order_details.status = 'SERVED'`.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 2, "notes": "Cập nhật trạng thái chi tiết"
    },

    # Epic 4: Menu & Catalog (Sprint 2)
    {
        "epic": "4. QUẢN LÝ DANH MỤC & THỰC ĐƠN (MENU)", "code": "US-21",
        "story": "Là Người quản lý, tôi muốn tạo mới danh mục món ăn (Category) để tôi phân loại thực đơn khoa học.",
        "points": 2, "priority": "Cao",
        "ac": "1. Có giao diện thêm danh mục món (Ví dụ: Món chính, Đồ uống, Tráng miệng...).\n2. Yêu cầu nhập Tên danh mục không trùng lặp.\n3. Cho phép cập nhật thông tin hoặc xóa danh mục (chỉ xóa khi danh mục không chứa sản phẩm).",
        "status": "Chưa bắt đầu", "dev": "Lộc", "tester": "Dũng", "sprint": 2, "notes": "Lưu bảng categories"
    },
    {
        "epic": "4. QUẢN LÝ DANH MỤC & THỰC ĐƠN (MENU)", "code": "US-22",
        "story": "Là Người quản lý, tôi muốn thêm món ăn mới vào thực đơn (Product) để tôi đa dạng hóa thực đơn phục vụ.",
        "points": 3, "priority": "Cao",
        "ac": "1. Form nhập liệu gồm: Tên món, Mô tả, Hình ảnh (upload lên server), Danh mục liên kết.\n2. Kiểm tra tính hợp lệ dữ liệu bắt buộc.\n3. Tự động đồng bộ món ăn mới lên danh sách món ăn hiển thị tại POS.",
        "status": "Chưa bắt đầu", "dev": "Lộc", "tester": "Dũng", "sprint": 2, "notes": "Lưu bảng products"
    },
    {
        "epic": "4. QUẢN LÝ DANH MỤC & THỰC ĐƠN (MENU)", "code": "US-23",
        "story": "Là Người quản lý, tôi muốn tạo các biến thể món ăn (Product Variants) để bán sản phẩm với nhiều kích cỡ và mức giá khác nhau.",
        "points": 3, "priority": "Cao",
        "ac": "1. Cho phép thêm nhiều biến thể dưới 1 sản phẩm gốc.\n2. Nhập thông tin: Tên biến thể (Size M, Size L...), Giá bán gốc, Giá vốn gốc, Mã SKU (duy nhất), Cờ Topping.\n3. Ràng buộc khóa ngoại chính xác đến sản phẩm gốc.",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 2, "notes": "Lưu product_variants"
    },
    {
        "epic": "4. QUẢN LÝ DANH MỤC & THỰC ĐƠN (MENU)", "code": "US-24",
        "story": "Là Đầu bếp, tôi muốn ngưng phục vụ món ăn tạm thời khi nguyên liệu chế biến món ăn đó bị hết trong ngày.",
        "points": 2, "priority": "Trung bình",
        "ac": "1. Cho phép chuyển đổi trạng thái `is_active = false` của sản phẩm hoặc biến thể tại trang quản trị.\n2. Món ăn bị tắt ngay lập tức ẩn khỏi màn hình POS của thu ngân.\n3. Tự động khôi phục hiển thị khi chuyển trạng thái thành `is_active = true`.",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 2, "notes": "Ẩn món tại POS"
    },
    {
        "epic": "4. QUẢN LÝ DANH MỤC & THỰC ĐƠN (MENU)", "code": "US-25",
        "story": "Là Người quản trị, tôi muốn xem lịch sử thay đổi giá bán của biến thể món ăn để giám sát biến động định giá sản phẩm.",
        "points": 2, "priority": "Thấp",
        "ac": "1. Mỗi khi cập nhật `price` của biến thể, tự động lưu một dòng nhật ký vào bảng `audit_logs`.\n2. Ghi rõ: Giá cũ, Giá mới, Người thực hiện thay đổi.\n3. Cho phép lọc lịch sử đổi giá theo thời gian.",
        "status": "Chưa bắt đầu", "dev": "Lộc", "tester": "Dũng", "sprint": 2, "notes": "Audit Log đổi giá"
    },

    # Epic 5: Inventory (Sprint 3)
    {
        "epic": "5. QUẢN LÝ KHO NGUYÊN LIỆU (INVENTORY)", "code": "US-26",
        "story": "Là Nhân viên kho, tôi muốn quản lý danh mục nguyên liệu thô để phân loại các nguyên vật liệu sử dụng cho bếp nấu.",
        "points": 3, "priority": "Cao",
        "ac": "1. Nhập thông tin: Mã nguyên liệu SKU (Ví dụ: MAT-001), Tên nguyên liệu, Đơn vị tính (kg, lít, lon, quả...), Ngưỡng tồn kho tối thiểu.\n2. Cho phép thêm mới, sửa thông tin nguyên vật liệu.",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 3, "notes": "Lưu inventory_items"
    },
    {
        "epic": "5. QUẢN LÝ KHO NGUYÊN LIỆU (INVENTORY)", "code": "US-27",
        "story": "Là Đầu bếp, tôi muốn thiết lập định mức nguyên liệu chế biến (Recipes) để hệ thống tự động trừ kho chi nhánh khi bán hàng.",
        "points": 5, "priority": "Cao",
        "ac": "1. Cho phép liên kết 1 biến thể sản phẩm với nhiều nguyên vật liệu thô.\n2. Nhập định lượng cần dùng (`quantity_needed` - Ví dụ: 0.15 kg thịt bò nạm cho 1 tô Phở Bò Size Lớn).\n3. Khi hóa đơn POS thanh toán thành công, hệ thống tự động trừ tồn kho nguyên liệu tương ứng trong kho chi nhánh.",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 3, "notes": "Định lượng Recipes"
    },
    {
        "epic": "5. QUẢN LÝ KHO NGUYÊN LIỆU (INVENTORY)", "code": "US-28",
        "story": "Là Nhân viên kho, tôi muốn ghi nhận xuất kho/nhập kho thủ công khi có hao hụt rơi vỡ hoặc nhập thêm nguyên liệu ngoài.",
        "points": 3, "priority": "Cao",
        "ac": "1. Có form tạo phiếu xuất/nhập kho thủ công.\n2. Chọn nguyên vật liệu, nhập số lượng thay đổi (+/-), lý do điều chỉnh.\n3. Tự động ghi nhật ký biến động vào bảng `inventory_logs` với loại STOCKIN hoặc STOCKOUT.",
        "status": "Chưa bắt đầu", "dev": "Lộc", "tester": "Dũng", "sprint": 3, "notes": "Điều chỉnh tồn kho"
    },
    {
        "epic": "5. QUẢN LÝ KHO NGUYÊN LIỆU (INVENTORY)", "code": "US-29",
        "story": "Là Quản lý chi nhánh, tôi muốn nhận cảnh báo tự động khi tồn kho nguyên liệu xuống thấp hơn ngưỡng tối thiểu để kịp thời mua thêm.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Hệ thống liên tục quét số lượng tồn kho thực tế trong `branch_inventory`.\n2. Nếu `quantity < reorder_point`, hiển thị biểu tượng cảnh báo màu vàng tại trang chủ Dashboard quản trị.\n3. Gợi ý lập đơn mua hàng PO cho nguyên vật liệu bị thiếu.",
        "status": "Chưa bắt đầu", "dev": "Lộc", "tester": "Dũng", "sprint": 3, "notes": "Cảnh báo reorder_point"
    },
    {
        "epic": "5. QUẢN LÝ KHO NGUYÊN LIỆU (INVENTORY)", "code": "US-30",
        "story": "Là Nhân viên kho, tôi muốn thực hiện kiểm kê kho định kỳ và điều chỉnh chênh lệch tồn tồn kho thực tế.",
        "points": 5, "priority": "Cao",
        "ac": "1. Cho phép tạo phiếu kiểm kê định kỳ.\n2. Nhập số lượng thực tế kiểm đếm tại kệ kho.\n3. Hệ thống tự động tính toán chênh lệch và cập nhật số lượng tồn kho về giá trị thực tế kiểm đếm, ghi log điều chỉnh.",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 3, "notes": "Kiểm kê & đối soát kho"
    },

    # Epic 6: Procurement (Sprint 3)
    {
        "epic": "6. THU MUA & NHÀ CUNG CẤP (PROCUREMENT)", "code": "US-31",
        "story": "Là Nhân viên mua hàng, tôi muốn quản lý thông tin nhà cung cấp (Supplier Profile) để liên hệ đặt mua hàng hóa ngoài.",
        "points": 2, "priority": "Cao",
        "ac": "1. Lưu trữ: Mã nhà cung cấp (Ví dụ: SUP-001), Tên nhà cung cấp, Email liên hệ, Số điện thoại, Địa chỉ.\n2. Ràng buộc không cho phép xóa nhà cung cấp đang có đơn hàng PO chưa hoàn thành.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 3, "notes": "Lưu bảng suppliers"
    },
    {
        "epic": "6. THU MUA & NHÀ CUNG CẤP (PROCUREMENT)", "code": "US-32",
        "story": "Là Nhân viên mua hàng, tôi muốn tạo đơn đặt hàng nhà cung cấp (Purchase Order - PO) để bổ sung nguyên vật liệu thô cho chi nhánh.",
        "points": 5, "priority": "Cao",
        "ac": "1. Cho phép tạo phiếu đặt hàng PO trạng thái 'DRAFT'.\n2. Chọn Supplier, Chi nhánh nhận hàng, Hạn chót giao hàng, danh sách nguyên liệu kèm số lượng và đơn giá nhập hợp đồng.\n3. Tính tổng số tiền đơn hàng PO tự động.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 3, "notes": "Khởi tạo purchase_orders"
    },
    {
        "epic": "6. THU MUA & NHÀ CUNG CẤP (PROCUREMENT)", "code": "US-33",
        "story": "Là Quản lý chi nhánh, tôi muốn xem xét phê duyệt đơn đặt hàng PO để nhân viên tiến hành gửi email đặt hàng cho Supplier.",
        "points": 3, "priority": "Cao",
        "ac": "1. Hiển thị danh sách các đơn hàng PO trạng thái DRAFT.\n2. Cho phép Quản lý nhấn 'Approve' (Chuyển trạng thái sang APPROVED) hoặc 'Reject' (Hủy đơn).\n3. Tự động gửi email thông báo đơn đặt hàng chứa danh sách nguyên vật liệu thô đến email nhà cung cấp.",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 3, "notes": "Phê duyệt & Gửi mail tự động"
    },
    {
        "epic": "6. THU MUA & NHÀ CUNG CẤP (PROCUREMENT)", "code": "US-34",
        "story": "Là Nhân viên kho, tôi muốn lập phiếu nhận hàng (Goods Receipt - GRN) khi nhà cung cấp giao nguyên liệu đến chi nhánh.",
        "points": 3, "priority": "Cao",
        "ac": "1. Cho phép chọn đơn hàng gốc PO đã được duyệt.\n2. Nhập số lượng thực tế nhận được (`quantity_received`) và số lượng bị lỗi từ chối nhận (`quantity_rejected`).\n3. Trạng thái đơn hàng PO cập nhật sang RECEIVED.",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 3, "notes": "Lưu phiếu goods_receipts"
    },
    {
        "epic": "6. THU MUA & NHÀ CUNG CẤP (PROCUREMENT)", "code": "US-35",
        "story": "Là Quản lý chi nhánh, tôi muốn thực hiện đối chiếu 3 bên (Invoice Matching) để xác nhận thanh toán tài chính cho nhà cung cấp.",
        "points": 5, "priority": "Trung bình",
        "ac": "1. Hệ thống tự động so khớp 3 thông tin: Đơn hàng ban đầu (PO) vs Phiếu thực nhận (GRN) vs Hóa đơn đòi tiền từ Supplier.\n2. Nếu sai lệch số lượng vượt quá 2%, hiển thị cảnh báo đỏ bắt buộc giải trình.\n3. Nếu khớp 100%, tự động cộng tồn kho vào chi nhánh và chuyển trạng thái PO thành hoàn tất thanh toán.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 3, "notes": "Quy trình Three-way Matching"
    },
    {
        "epic": "6. THU MUA & NHÀ CUNG CẤP (PROCUREMENT)", "code": "US-36",
        "story": "Là Nhân viên mua hàng, tôi muốn nhận cảnh báo đơn hàng PO bị giao trễ hạn để tôi liên hệ hối thúc nhà cung cấp.",
        "points": 2, "priority": "Thấp",
        "ac": "1. Hệ thống quét các đơn PO trạng thái APPROVED chưa hoàn thành nhận hàng.\n2. Nếu ngày hiện tại vượt quá `delivery_deadline`, đổi trạng thái hiển thị của PO thành DELAYED trên trang danh sách đơn PO.",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 3, "notes": "Cảnh báo giao hàng trễ"
    },

    # Epic 7: HR & Payroll (Sprint 4)
    {
        "epic": "7. NHÂN SỰ, CHẤM CÔNG & TÍNH LƯƠNG (HR)", "code": "US-37",
        "story": "Là Nhân viên nhân sự (HR), tôi muốn quản lý thông tin hồ sơ nhân viên để lưu trữ tập trung dữ liệu nhân sự của hệ thống.",
        "points": 3, "priority": "Cao",
        "ac": "1. Lưu trữ: Mã nhân viên, Liên kết tài khoản hệ thống (`user_id`), Bộ phận làm việc, Chức danh, Ngày vào làm, Mức lương cơ bản, Hình thức tính lương (Fixed hoặc Hourly), Chi nhánh trực thuộc.\n2. Cho phép thêm mới nhân viên, sửa hồ sơ, chuyển chi nhánh hoạt động.",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 4, "notes": "Lưu bảng employees"
    },
    {
        "epic": "7. NHÂN SỰ, CHẤM CÔNG & TÍNH LƯƠNG (HR)", "code": "US-38",
        "story": "Là Nhân viên nhân sự (HR), tôi muốn thiết lập mẫu ca làm việc chuẩn để áp dụng phân ca nhanh chóng hàng tuần.",
        "points": 2, "priority": "Cao",
        "ac": "1. Cho phép tạo các ca làm việc mẫu (Ví dụ: Ca sáng 8:00 - 16:00, Ca gãy 10:00 - 14:00...).\n2. Nhập: Tên ca, Giờ bắt đầu, Giờ kết thúc, Số giờ công quy đổi.\n3. Cho phép chỉnh sửa thông tin ca mẫu.",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 4, "notes": "Lưu shift_templates"
    },
    {
        "epic": "7. NHÂN SỰ, CHẤM CÔNG & TÍNH LƯƠNG (HR)", "code": "US-39",
        "story": "Là Nhân viên nhân sự (HR), tôi muốn phân lịch làm việc tuần cho nhân viên để nhân viên biết chính xác ca trực của mình.",
        "points": 3, "priority": "Cao",
        "ac": "1. Có giao diện bảng phân ca theo tuần (Thứ 2 đến Chủ nhật).\n2. Cho phép gán nhanh Ca mẫu cho nhân viên vào ngày cụ thể.\n3. Ngăn chặn gán trùng 2 ca chồng chéo giờ làm trong cùng 1 ngày cho 1 nhân viên.",
        "status": "Chưa bắt đầu", "dev": "Lộc", "tester": "Dũng", "sprint": 4, "notes": "Lưu employee_shift_assignments"
    },
    {
        "epic": "7. NHÂN SỰ, CHẤM CÔNG & TÍNH LƯƠNG (HR)", "code": "US-40",
        "story": "Là Nhân viên phục vụ, tôi muốn thực hiện chấm công vào/ra (Clock In / Clock Out) để ghi nhận thời gian làm việc thực tế.",
        "points": 3, "priority": "Cao",
        "ac": "1. Có nút Clock In/Out nhanh trên giao diện nhân viên.\n2. Clock In: ghi nhận thời gian bắt đầu trực, tự động so khớp ca gán để đánh dấu đi trễ (is_late).\n3. Clock Out: ghi nhận thời gian ra ca, tính tổng số giờ công làm thực tế (hours_worked), tự động đánh dấu về sớm.",
        "status": "Chưa bắt đầu", "dev": "Lộc", "tester": "Dũng", "sprint": 4, "notes": "Lưu employee_attendances"
    },
    {
        "epic": "7. NHÂN SỰ, CHẤM CÔNG & TÍNH LƯƠNG (HR)", "code": "US-41",
        "story": "Là Nhân viên phục vụ quên chấm công, tôi muốn gửi đơn báo quên chấm công (Forgot Clock Request) để xin Quản lý phê duyệt bổ sung công ca làm.",
        "points": 2, "priority": "Trung bình",
        "ac": "1. Có form tạo yêu cầu quên chấm công.\n2. Chọn ngày quên, Giờ đề xuất bổ sung, Loại quên (Clock In hoặc Clock Out), Lý do giải trình.\n3. Trạng thái mặc định của đơn là PENDING.",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 4, "notes": "Tạo forgot_clock_requests"
    },
    {
        "epic": "7. NHÂN SỰ, CHẤM CÔNG & TÍNH LƯƠNG (HR)", "code": "US-42",
        "story": "Là Nhân viên phục vụ, tôi muốn gửi đơn xin nghỉ phép trực tuyến (Leave Request) để xin Quản lý duyệt nghỉ phép có lương hoặc không lương.",
        "points": 2, "priority": "Trung bình",
        "ac": "1. Có form tạo đơn nghỉ phép.\n2. Chọn Ngày bắt đầu, Ngày kết thúc, Loại nghỉ phép (Phép năm ANNUAL, Nghỉ ốm SICK, Nghỉ không lương UNPAID), Lý do xin nghỉ.\n3. Tự động kiểm tra số ngày phép còn lại của nhân viên (nếu chọn ANNUAL).",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 4, "notes": "Tạo leave_requests"
    },
    {
        "epic": "7. NHÂN SỰ, CHẤM CÔNG & TÍNH LƯƠNG (HR)", "code": "US-43",
        "story": "Là Quản lý chi nhánh, tôi muốn phê duyệt hoặc từ chối các đơn từ của nhân viên (Leave / Forgot Clock) để tôi chốt công nhanh chóng.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Có màn hình duyệt đơn tập trung của chi nhánh.\n2. Quản lý xem chi tiết giải trình và nhấn APPROVED hoặc REJECTED.\n3. Nếu duyệt APPROVED Forgot Clock, hệ thống tự động chèn/cập nhật giờ công vào bảng `employee_attendances` tương ứng.",
        "status": "Chưa bắt đầu", "dev": "Lộc", "tester": "Dũng", "sprint": 4, "notes": "Tự động bù công sau khi duyệt"
    },
    {
        "epic": "7. NHÂN SỰ, CHẤM CÔNG & TÍNH LƯƠNG (HR)", "code": "US-44",
        "story": "Là Nhân viên nhân sự (HR), tôi muốn chạy tính lương tháng tự động (Payroll Run) để tính toán thu nhập thực tế chính xác cho nhân viên.",
        "points": 5, "priority": "Cao",
        "ac": "1. Cho phép kích hoạt chạy bảng lương cho 1 tháng (Định dạng YYYY-MM).\n2. Hệ thống quét toàn bộ ngày công của nhân viên: Tính lương cứng cố định hoặc nhân giờ công nhân đơn giá giờ.\n3. Cộng các khoản allowances và trừ các khoản phạt đi trễ/đơn nghỉ không lương (deductions), lưu kết quả vào bảng `payroll_entries`.",
        "status": "Chưa bắt đầu", "dev": "Lộc", "tester": "Dũng", "sprint": 4, "notes": "Chạy bảng lương tháng tự động"
    },

    # Epic 8: Loyalty - Redesigned Epic (Sprint 4)
    {
        "epic": "8. KHÁCH HÀNG THÂN THIẾT & TÍCH ĐIỂM (LOYALTY)", "code": "US-45",
        "story": "Là Thu ngân, tôi muốn đăng ký hồ sơ hội viên mới tại POS lấy Số điện thoại (SĐT) làm định danh khóa chính duy nhất để tối ưu hóa tra cứu.",
        "points": 5, "priority": "Trung bình",
        "ac": "1. Form nhập liệu gồm: SĐT (phone), Họ tên (name), Ngày sinh (birth_date).\n2. Kiểm tra tính hợp lệ: SĐT bắt buộc là duy nhất, không sử dụng cột surrogate ID tự tăng.\n3. Khởi tạo hạng thẻ mặc định là 'Bronze', tích lũy chi tiêu `total_spent = 0.0`.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 4, "notes": "Redesign: Phone as PK"
    },
    {
        "epic": "8. KHÁCH HÀNG THÂN THIẾT & TÍCH ĐIỂM (LOYALTY)", "code": "US-46",
        "story": "Là Hệ thống, tôi muốn tự động tích lũy chi tiêu và cộng điểm thưởng khi hội viên thanh toán hóa đơn để ghi nhận giá trị mua sắm.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Khi hóa đơn POS gán SĐT hội viên đổi sang PAID, hệ thống kích hoạt Loyalty Service.\n2. Cộng dồn số tiền thanh toán thực tế vào cột `customers.total_spent`.\n3. Cộng điểm thưởng `loyalty_points` tương đương 1% tổng số tiền hóa đơn, lưu lịch sử loại EARNED.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 4, "notes": "Tự động tích lũy sau thanh toán"
    },
    {
        "epic": "8. KHÁCH HÀNG THÂN THIẾT & TÍCH ĐIỂM (LOYALTY)", "code": "US-47",
        "story": "Là Hệ thống, tôi muốn tự động nâng hạng thẻ thành viên (Membership Tiering) dựa trên chi tiêu tích lũy để tri ân đặc quyền cho khách.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Sau khi cộng doanh số, so sánh `total_spent` để nâng hạng: Silver >= 5M VND, Gold >= 15M VND, Platinum >= 50M VND.\n2. Tự động cập nhật `membership_tier` trên bảng khách hàng.\n3. Mức chiết khấu mặc định của hạng thẻ tự động áp dụng cho hóa đơn kế tiếp (Silver: 5%, Gold: 10%, Platinum: 15%).",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 4, "notes": "Tự động xếp hạng nâng thẻ"
    },
    {
        "epic": "8. KHÁCH HÀNG THÂN THIẾT & TÍCH ĐIỂM (LOYALTY)", "code": "US-48",
        "story": "Là Thu ngân, tôi muốn quy đổi điểm tích lũy của khách hàng thành tiền trừ trực tiếp vào hóa đơn khi có yêu cầu của khách.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Cho phép nhập số điểm muốn quy đổi tại giỏ hàng POS (Tỷ lệ quy đổi: 1 điểm = 1,000 VNĐ).\n2. Khấu trừ trực tiếp số tiền tương ứng vào tổng tiền hóa đơn thanh toán.\n3. Trừ điểm thưởng của khách trên DB và lưu vết giao dịch loại REDEEMED.",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 4, "notes": "Quy đổi điểm thưởng trực tiếp"
    },
    {
        "epic": "8. KHÁCH HÀNG THÂN THIẾT & TÍCH ĐIỂM (LOYALTY)", "code": "US-49",
        "story": "Là Hệ thống, tôi muốn tự động phát hành voucher giảm giá mừng sinh nhật hội viên để thúc đẩy doanh thu chăm sóc khách.",
        "points": 2, "priority": "Thấp",
        "ac": "1. Hệ thống tự động quét ngày sinh nhật của khách hàng hàng ngày.\n2. Nếu trùng ngày sinh, tự động tạo mã coupon mừng sinh nhật thời hạn sử dụng 7 ngày.\n3. Gửi thông báo coupon về Cổng thông tin khách hàng.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 4, "notes": "Tặng voucher sinh nhật tự động"
    },
    {
        "epic": "8. KHÁCH HÀNG THÂN THIẾT & TÍCH ĐIỂM (LOYALTY)", "code": "US-50",
        "story": "Là Khách hàng, tôi muốn tự truy cập cổng Customer Portal bằng số điện thoại để kiểm tra điểm thưởng và hạng thẻ của mình.",
        "points": 2, "priority": "Thấp",
        "ac": "1. Khách hàng truy cập đường dẫn `/customer-portal/{phone}`.\n2. Hiển thị thông tin: Họ tên, Hạng thẻ hiện tại, Số điểm tích lũy, Lịch sử giao dịch tích điểm/tiêu điểm.\n3. Giao diện tối ưu hiển thị trên các thiết bị di động cá nhân.",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 4, "notes": "Customer Portal hiển thị di động"
    },

    # Epic 9: Promotion (Sprint 5)
    {
        "epic": "9. CHƯƠNG TRÌNH KHUYẾN MÃI (PROMOTION)", "code": "US-51",
        "story": "Là Quản lý chi nhánh, tôi muốn cấu hình chương trình Khuyến mãi giảm giá theo món để áp dụng giảm trực tiếp giá món ăn.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Nhập thông tin khuyến mãi: Tên chương trình, Mã coupon, Giá trị giảm (số tiền cố định hoặc tỷ lệ %), Thời gian hiệu lực.\n2. Cho phép chọn món ăn áp dụng trong thực đơn.\n3. Món ăn được hiển thị giá đã giảm trên màn hình bán hàng POS.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 5, "notes": "Khuyến mãi giảm giá theo món"
    },
    {
        "epic": "9. CHƯƠNG TRÌNH KHUYẾN MÃI (PROMOTION)", "code": "US-52",
        "story": "Là Quản lý chi nhánh, tôi muốn cấu hình chương trình Mua 1 Tặng 1 (Buy 1 Get 1 - B1G1) để đẩy nhanh doanh số tiêu thụ các món mới.",
        "points": 5, "priority": "Trung bình",
        "ac": "1. Chọn sản phẩm kích hoạt (triggerProductId) và sản phẩm quà tặng (rewardProductId).\n2. Khi thu ngân thêm sản phẩm kích hoạt vào giỏ hàng POS, Promotion Engine tự động thêm món tặng giá 0 VNĐ vào giỏ.\n3. Tự động xóa món tặng nếu sản phẩm kích hoạt bị xóa khỏi giỏ.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 5, "notes": "Khuyến mãi Buy-1-Get-1"
    },
    {
        "epic": "9. CHƯƠNG TRÌNH KHUYẾN MÃI (PROMOTION)", "code": "US-53",
        "story": "Là Thu ngân, tôi muốn nhập mã giảm giá Coupon tại màn hình POS để chiết khấu hóa đơn cho khách có voucher.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Có ô nhập mã Coupon tại giao diện checkout POS.\n2. Kiểm tra tính hợp lệ: Mã có tồn tại? Đơn hàng đạt giá trị tối thiểu? Đơn hàng nằm trong thời gian áp dụng? Giới hạn lượt dùng còn không?\n3. Áp dụng mức giảm giá vào tổng bill và lưu lịch sử dùng coupon vào bảng `promotion_usage`.",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 5, "notes": "Áp dụng coupon thủ công"
    },
    {
        "epic": "9. CHƯƠNG TRÌNH KHUYẾN MÃI (PROMOTION)", "code": "US-54",
        "story": "Là Hệ thống, tôi muốn tự động áp dụng chương trình khuyến mãi có lợi nhất cho khách hàng để tăng trải nghiệm hài lòng mua sắm.",
        "points": 5, "priority": "Trung bình",
        "ac": "1. Khi thanh toán, Promotion Engine tự động duyệt toàn bộ coupon khả dụng trong DB.\n2. So sánh số tiền được giảm của từng coupon.\n3. Tự động áp dụng coupon giảm giá nhiều tiền nhất (tối ưu nhất) vào hóa đơn mà khách hàng không cần nhập mã.",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 5, "notes": "Thuật toán áp dụng Optimal Promo"
    },

    # Epic 10: Multi-Branch - Redesigned Epic (Sprint 5)
    {
        "epic": "10. VẬN HÀNH ĐA CHI NHÁNH (MULTI-BRANCH)", "code": "US-55",
        "story": "Là Chủ chuỗi (Admin), tôi muốn quản lý danh sách các chi nhánh cửa hàng để tôi mở rộng hoặc theo dõi toàn bộ cơ sở kinh doanh.",
        "points": 3, "priority": "Cao",
        "ac": "1. Giao diện quản lý chi nhánh gồm: Mã chi nhánh (branch_id), Tên chi nhánh, Địa chỉ, Số điện thoại, trạng thái hoạt động.\n2. Cho phép thêm mới chi nhánh, tạm ngưng hoạt động (`is_active = false`) hoặc mở lại chi nhánh.",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 5, "notes": "Quản lý bảng branches"
    },
    {
        "epic": "10. VẬN HÀNH ĐA CHI NHÁNH (MULTI-BRANCH)", "code": "US-56",
        "story": "Là Chủ chuỗi (Admin), tôi muốn hệ thống phân tách dữ liệu hoạt động theo chi nhánh để bảo mật thông tin và tránh nhầm lẫn.",
        "points": 5, "priority": "Cao",
        "ac": "1. Tài khoản Cashier/Manager được gán cố định `branch_id`.\n2. Spring Security chặn yêu cầu và tự động lọc dữ liệu qua JPA: Chỉ cho phép xem/thao tác kho, bàn ăn, nhân sự, hóa đơn thuộc chi nhánh mình trực thuộc.\n3. Tài khoản Super Admin (`branch_id = NULL`) có quyền xem toàn chuỗi.",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 5, "notes": "Cơ chế Logical Isolation"
    },
    {
        "epic": "10. VẬN HÀNH ĐA CHI NHÁNH (MULTI-BRANCH)", "code": "US-57",
        "story": "Là Người quản lý, tôi muốn điều chỉnh giá bán và trạng thái món ăn riêng biệt theo chi nhánh để bù đắp chênh lệch chi phí vận hành vùng miền.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Có giao diện cấu hình giá thực đơn riêng cho chi nhánh.\n2. Cho phép điền `custom_price` của từng biến thể món ăn riêng biệt.\n3. Cho phép bật/tắt phục vụ món ăn dựa trên nguồn nguyên liệu tại chỗ, lưu vào bảng `branch_product_prices`.",
        "status": "Chưa bắt đầu", "dev": "Lộc", "tester": "Dũng", "sprint": 5, "notes": "Giá bán riêng chi nhánh"
    },
    {
        "epic": "10. VẬN HÀNH ĐA CHI NHÁNH (MULTI-BRANCH)", "code": "US-58",
        "story": "Là Quản lý chi nhánh con, tôi muốn tạo phiếu đặt hàng bổ sung nguyên vật liệu nội bộ (IPO) gửi lên Kho tổng khi chi nhánh sắp hết hàng.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Tự động gợi ý tạo đơn khi tồn kho dưới ngưỡng `reorder_point`.\n2. Cho phép tạo đơn hàng nội bộ IPO chọn gửi đến chi nhánh được định nghĩa là Kho tổng (`is_warehouse = true`).\n3. Trạng thái phiếu đặt nội bộ mặc định là SUBMITTED.",
        "status": "Chưa bắt đầu", "dev": "Lộc", "tester": "Dũng", "sprint": 5, "notes": "Đơn hàng nội bộ IPO"
    },
    {
        "epic": "10. VẬN HÀNH ĐA CHI NHÁNH (MULTI-BRANCH)", "code": "US-59",
        "story": "Là Quản lý chi nhánh, tôi muốn thực hiện quy trình chuyển kho nội bộ và đối soát hao hụt để luân chuyển nguyên vật liệu trong chuỗi.",
        "points": 5, "priority": "Trung bình",
        "ac": "1. Xuất hàng (Ship): Giảm tồn kho ở kho nguồn, đổi trạng thái phiếu sang SHIPPED, hàng đi vào trạng thái Đang vận chuyển.\n2. Nhận hàng (Receive): Kho đích kiểm đếm nhập số lượng thực nhận, tự động tính chênh lệch hao hụt rơi vỡ (`loss_quantity = shipped - received`) và lý do.\n3. Cộng kho đích theo số lượng thực nhận, lưu lịch sử logs.",
        "status": "Chưa bắt đầu", "dev": "Dũng", "tester": "Lộc", "sprint": 5, "notes": "Chuyển kho & đối soát hao hụt"
    },

    # Epic 11: Analytics & AI (Sprint 6)
    {
        "epic": "11. BÁO CÁO & PHÂN TÍCH (ANALYTICS)", "code": "US-60",
        "story": "Là Quản lý chi nhánh, tôi muốn xem báo cáo doanh thu theo chi nhánh trực quan theo ngày, tuần, tháng để đánh giá tình hình kinh doanh.",
        "points": 5, "priority": "Cao",
        "ac": "1. Có trang Dashboard hiển thị biểu đồ đường cột doanh thu.\n2. Hỗ trợ lọc doanh thu theo Chi nhánh, Phương thức thanh toán (Tiền mặt/VNPay), Thời gian.\n3. Lấy dữ liệu từ bảng orders có trạng thái SERVED.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 6, "notes": "Dashboard báo cáo doanh số"
    },
    {
        "epic": "11. BÁO CÁO & PHÂN TÍCH (ANALYTICS)", "code": "US-61",
        "story": "Là Quản lý, tôi muốn xuất báo cáo doanh thu và báo cáo xuất nhập kho ra tệp Excel để phục vụ đối soát kế toán.",
        "points": 3, "priority": "Cao",
        "ac": "1. Có nút 'Xuất Excel' trên các trang báo cáo doanh số và báo cáo tồn kho.\n2. Tự động sinh tệp Excel chứa dữ liệu thô đã được định dạng cột rõ ràng.\n3. Cho phép tải tệp trực tiếp về máy tính cá nhân.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 6, "notes": "Tải báo cáo định dạng Excel"
    },
    {
        "epic": "11. BÁO CÁO & PHÂN TÍCH (ANALYTICS)", "code": "US-62",
        "story": "Là Người quản lý, tôi muốn xem thống kê hiệu quả sử dụng của các chương trình khuyến mãi để điều chỉnh ngân sách tiếp thị.",
        "points": 3, "priority": "Trung bình",
        "ac": "1. Hiển thị báo cáo số lượng coupon đã sử dụng, tổng số tiền đã giảm giá theo từng chương trình.\n2. Tính toán tỷ lệ chuyển đổi của các chiến dịch.\n3. Lọc dữ liệu dựa trên bảng promotion_usage.",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 6, "notes": "Báo cáo hiệu quả khuyến mãi"
    },
    {
        "epic": "11. BÁO CÁO & PHÂN TÍCH (ANALYTICS)", "code": "US-63",
        "story": "Là Người quản lý, tôi muốn xem thống kê tỷ lệ quay lại của khách hàng thành VIP để đánh giá mức độ trung thành của hội viên.",
        "points": 2, "priority": "Thấp",
        "ac": "1. Báo cáo phân tích số lượng khách hàng theo từng phân hạng thẻ (Bronze, Silver, Gold, Platinum).\n2. Thống kê tần suất ăn uống trung bình của khách hàng thành viên có gán SĐT.\n3. Hiển thị biểu đồ phân bố trực quan.",
        "status": "Chưa bắt đầu", "dev": "Thuận", "tester": "Kha", "sprint": 6, "notes": "Báo cáo phân tích hội viên VIP"
    },
    {
        "epic": "11. BÁO CÁO & PHÂN TÍCH (ANALYTICS)", "code": "US-64",
        "story": "Là Chủ chuỗi (Admin), tôi muốn trò chuyện trực tiếp với Trợ lý AI Assistant để phân tích nhanh doanh số và đề xuất giải pháp tối ưu hoạt động chuỗi.",
        "points": 5, "priority": "Trung bình",
        "ac": "1. Có giao diện chatbox AI tích hợp ngay trên Dashboard quản trị.\n2. Cho phép nhập câu hỏi phân tích (Ví dụ: 'Món ăn nào bán chạy nhất chi nhánh quận 1 tháng này?').\n3. Trợ lý AI kết nối an toàn với Gemini API, truy vấn cơ sở dữ liệu bán hàng và trả về câu trả lời phân tích dưới 3 giây.",
        "status": "Chưa bắt đầu", "dev": "Kha", "tester": "Thuận", "sprint": 6, "notes": "Tích hợp Gemini AI Service"
    }
]

def generate_xlsx(xlsx_path):
    print(f"Generating beautifully styled Excel Backlog to {xlsx_path}...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RMS Project Backlog"
    
    # Freeze the header row and first two columns
    ws.freeze_panes = "C2"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styles Definition
    font_family = "Segoe UI"
    
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # FPT Navy Blue
    
    row_font = Font(name=font_family, size=10, color="1F2937")
    code_font = Font(name=font_family, size=10, bold=True, color="1E3A8A")
    epic_font = Font(name=font_family, size=10, bold=True, color="374151")
    
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border_side = Side(style='thin', color='E5E7EB')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    headers = [
        "Epic / Phân Hệ", 
        "Mã US", 
        "Tên Chức Năng (User Story chuẩn Agile)", 
        "Story Points", 
        "Độ Ưu Tiên", 
        "Acceptance Criteria (Tiêu Chí Nghiệm Thu Chi Tiết)", 
        "Trạng Thái", 
        "Người Phụ Trách (Dev)", 
        "Người Phụ Trách (Tester)", 
        "Sprint Dự Kiến", 
        "Ghi Chú"
    ]
    
    # Write Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    
    # Write Data
    for row_idx, item in enumerate(backlog_data, 2):
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        
        row_cells = [
            ws.cell(row=row_idx, column=1, value=item["epic"]),
            ws.cell(row=row_idx, column=2, value=item["code"]),
            ws.cell(row=row_idx, column=3, value=item["story"]),
            ws.cell(row=row_idx, column=4, value=item["points"]),
            ws.cell(row=row_idx, column=5, value=item["priority"]),
            ws.cell(row=row_idx, column=6, value=item["ac"]),
            ws.cell(row=row_idx, column=7, value=item["status"]),
            ws.cell(row=row_idx, column=8, value=item["dev"]),
            ws.cell(row=row_idx, column=9, value=item["tester"]),
            ws.cell(row=row_idx, column=10, value=item["sprint"]),
            ws.cell(row=row_idx, column=11, value=item["notes"])
        ]
        
        for col_idx, cell in enumerate(row_cells, 1):
            cell.font = row_font
            cell.fill = row_fill
            cell.border = thin_border
            
            # Alignments
            if col_idx in [1]:
                cell.alignment = align_left
                cell.font = epic_font
            elif col_idx in [2]:
                cell.alignment = align_center
                cell.font = code_font
            elif col_idx in [4, 5, 7, 8, 9, 10]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    # Auto-fit Column Widths with comfortable padding
    column_widths = {
        1: 30, # Epic
        2: 10, # Code
        3: 50, # Story
        4: 12, # Points
        5: 12, # Priority
        6: 60, # AC
        7: 15, # Status
        8: 20, # Dev
        9: 20, # Tester
        10: 15, # Sprint
        11: 30  # Notes
    }
    
    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        
    # Enable filtering on all columns
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(backlog_data) + 1}"
    
    # Save Workbook
    wb.save(xlsx_path)
    print(f"Successfully generated styled Excel file: {xlsx_path}!")

def generate_csv(csv_path):
    print(f"Generating CSV Backlog to {csv_path}...")
    headers = [
        "Epic / Phân Hệ", 
        "Mã US", 
        "Tên Chức Năng (User Story chuẩn Agile)", 
        "Story Points", 
        "Độ Ưu Tiên", 
        "Acceptance Criteria (Tiêu Chí Nghiệm Thu Chi Tiết)", 
        "Trạng Thái", 
        "Người Phụ Trách (Dev)", 
        "Người Phụ Trách (Tester)", 
        "Sprint Dự Kiến", 
        "Ghi Chú"
    ]
    
    with open(csv_path, mode="w", newline="", encoding="utf-8-sig") as f: # utf-8-sig makes it Excel-friendly
        writer = csv.writer(f)
        writer.writerow(headers)
        
        for item in backlog_data:
            writer.writerow([
                item["epic"],
                item["code"],
                item["story"],
                item["points"],
                item["priority"],
                item["ac"],
                item["status"],
                item["dev"],
                item["tester"],
                item["sprint"],
                item["notes"]
            ])
            
    print(f"Successfully generated CSV file: {csv_path}!")

if __name__ == "__main__":
    xlsx_path = "d:/swp/technical_docs/Project_Backlog_SWP391.xlsx"
    csv_path = "d:/swp/technical_docs/Project_Backlog_SWP391.csv"
    
    generate_xlsx(xlsx_path)
    generate_csv(csv_path)
    print("\n--- ALL BACKLOG ASSETS SUCCESSFULLY GENERATED ---")
